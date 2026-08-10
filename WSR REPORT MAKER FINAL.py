
"""
Full Automation Script v12
==========================

CHANGES FROM v11:

  1. REMOVED tabs: "Unique" and "DistinctUnique" — no longer generated.

  2. ALL pivot tables (WTBP, VolumeAnalysis, MonthlyVolume) now:
       • Source directly from "Original with Dups" (raw, no deduplication step)
       • Values = DISTINCT COUNT of the Number column
         (i.e. how many unique Number values appear in each cell's group/week/month)

  3. Charts are built from the same distinct-count pivot tables.

  4. Traffic-light colouring (avg-based ±20%) still applies to all value cells.

  5. Closure tabs unchanged — still filter by Assignment Group, dedup on Task Number,
     and summarise by Assigned To × Month.

  6. MANAGED_TABS updated — Unique and DistinctUnique removed.

KEY BEHAVIOUR:
  For every pivot cell the script computes:
    len( unique Number values in that Priority/Group × Week/Month slice )
  This is equivalent to Excel's COUNTDISTINCT / DISTINCTCOUNT.
"""

import sys, os, datetime, warnings, traceback, io, tempfile

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
import openpyxl.chart.label

warnings.filterwarnings("ignore")

MASTER_DATA_PATH = r""
RAW_DATA_PATH = r""

PRIORITY_ORDER = ["1 - Critical", "2 - High", "3 - Moderate", "4 - Low", "5 - Planning"]
PRIORITY_LABELS = {
    "1 - Critical": "1 Critical",
    "2 - High": "2 High",
    "3 - Moderate": "3 Moderate",
    "4 - Low": "4 Low",
    "5 - Planning": "5 Planning",
}

DARK_BLUE = "1E3A5F"
MID_BLUE = "2E86AB"
WHITE = "FFFFFF"
ALT_PAT = PatternFill("solid", fgColor="EEF2FA")
HEADER_FILL = PatternFill("solid", fgColor=DARK_BLUE)
TOTAL_FILL = PatternFill("solid", fgColor=MID_BLUE)

# Traffic-light fills (avg-based)
RED_FILL = PatternFill("solid", fgColor="FF4C4C")
AMBER_FILL = PatternFill("solid", fgColor="FFB347")
GREEN_FILL = PatternFill("solid", fgColor="70C95E")

# ── Closure group config ──────────────────────────────────────────────────────
CLOSURE_GROUPS = [
    ("FTS-GT-L3-Datalake", "ClosureGCP", "ClosureGCP_Table"),
    ("FTS-GT-L3-Looker", "ClosureLooker", "ClosureLooker_Table"),
    ("FTS-GT-L3-SAP BI", "ClosureBW", "ClosureBW_Table"),
]

# Unique & DistinctUnique removed
MANAGED_TABS = [
    "Original with Dups",
    "WTBP",
    "VolumeAnalysis",
    "MonthlyVolume",
    "Charts",
    "ClosureGCP",
    "ClosureLooker",
    "ClosureBW",
    "ClosureGCP_Table",
    "ClosureLooker_Table",
    "ClosureBW_Table",
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def fmt_week(d):
    if d is None: return ""
    if isinstance(d, datetime.date):
        return d.strftime("%d-%b").lstrip("0")
    return str(d)

def fmt_month(d):
    if d is None: return ""
    if isinstance(d, datetime.date):
        return d.strftime("%b-%Y")
    return str(d)

def _side(): return Side(style="thin", color="BFBFBF")
def _border(): s = _side(); return Border(left=s, right=s, top=s, bottom=s)

def _cell(ws, r, c, v, bold=False, fc="000000", fill=None,
          align="center", wrap=False, sz=10):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name="Arial", size=sz, bold=bold, color=fc)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = _border()
    if fill: cell.fill = fill
    return cell

def hcell(ws, r, c, v, fill=None):
    return _cell(ws, r, c, v, bold=True, fc=WHITE,
                 fill=fill or HEADER_FILL, align="center", wrap=True)

def dcell(ws, r, c, v, fill=None, align="center"):
    return _cell(ws, r, c, v, fill=fill, align=align)

def autofit(ws, mn=8, mx=40):
    for col in ws.columns:
        ltr = get_column_letter(col[0].column)
        w = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[ltr].width = min(max(w + 3, mn), mx)

def find_col(df, names):
    lower = {c.lower().strip(): c for c in df.columns}
    for n in names:
        if n.lower() in lower: return lower[n.lower()]
    return None

def _read_upload(uploader):
    val = uploader.value
    if not val: return None, None
    if isinstance(val, dict):
        name = list(val.keys())[0]
        content = bytes(val[name]["content"])
    else:
        name = val[0]["name"]
        content = bytes(val[0]["content"])
    return name, content


# ── Week & month detection ────────────────────────────────────────────────────

def get_weeks_and_months_from_bytes(content: bytes):
    df = pd.read_excel(io.BytesIO(content), sheet_name=0)
    sc = find_col(df, ["start", "start date", "start_date", "created"])
    if sc is None:
        return [], [], df
    df[sc] = pd.to_datetime(df[sc], errors="coerce")
    valid = df[sc].dropna()
    weeks = sorted(valid.apply(
        lambda d: (d - pd.Timedelta(days=d.weekday())).date()
    ).unique())
    months = sorted(valid.apply(
        lambda d: d.date().replace(day=1)
    ).unique())
    return weeks, months, df


# ── Traffic-light helpers ─────────────────────────────────────────────────────

def _traffic_fill_avg(val, avg, tolerance=0.20):
    if avg == 0:
        return RED_FILL if val == 0 else GREEN_FILL
    lo = avg * (1 - tolerance)
    hi = avg * (1 + tolerance)
    if val > hi: return GREEN_FILL
    elif val >= lo: return AMBER_FILL
    else: return RED_FILL

def _tl_dcell(ws, r, c, val, avg):
    fill = _traffic_fill_avg(val, avg)
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(name="Arial", size=10, color="000000")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _border()
    cell.fill = fill
    return cell


# ── Distinct-count pivot builder ──────────────────────────────────────────────

def _distinct_count_pivot(df, index_col, col_col, number_col):
    """
    Returns a DataFrame where each cell = number of DISTINCT values in
    `number_col` for that (index_col, col_col) combination.
    Equivalent to COUNTDISTINCT / DISTINCTCOUNT in Excel / DAX.
    """
    if df.empty or number_col is None:
        return pd.DataFrame()
    return (
        df.groupby([index_col, col_col])[number_col]
        .nunique()
        .unstack(fill_value=0)
    )


# ── Combo chart builder ───────────────────────────────────────────────────────

def _combo(ws_d, d_start, d_end, cat_c, v_start, v_end, tot_c,
           title, w=24, h=14):
    bar = BarChart(); bar.type = "col"; bar.grouping = "clustered"
    bar.title = title; bar.style = 10; bar.width = w; bar.height = h
    bar.y_axis.title = "Distinct Number Count"; bar.y_axis.numFmt = "0"
    bar.y_axis.majorGridlines = None
    for vc in range(v_start, v_end + 1):
        bar.add_data(Reference(ws_d, min_col=vc, min_row=d_start, max_row=d_end),
                     titles_from_data=True)
    cats = Reference(ws_d, min_col=cat_c, min_row=d_start + 1, max_row=d_end)
    bar.set_categories(cats)
    for s in bar.series:
        s.dLbls = openpyxl.chart.label.DataLabelList()
        s.dLbls.showVal = True; s.dLbls.showLegendKey = False
        s.dLbls.showCatName = False; s.dLbls.showSerName = False
    line = LineChart()
    line.add_data(Reference(ws_d, min_col=tot_c, min_row=d_start, max_row=d_end),
                  titles_from_data=True)
    line.set_categories(cats)
    ls = line.series[0]
    ls.marker.symbol = "circle"; ls.marker.size = 7
    ls.graphicalProperties.line.solidFill = "FF0000"
    ls.graphicalProperties.line.width = 25000
    ls.dLbls = openpyxl.chart.label.DataLabelList(); ls.dLbls.showVal = True
    bar += line
    return bar


# ── Clear sheet ───────────────────────────────────────────────────────────────

def _clear_sheet(ws):
    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue
            cell.value = None; cell.font = Font(); cell.fill = PatternFill()
            cell.border = Border(); cell.alignment = Alignment()
            cell.number_format = "General"
    ws.row_dimensions.clear(); ws.column_dimensions.clear()
    ws.freeze_panes = None


# ── Generic pivot table writer ────────────────────────────────────────────────

def _write_pivot_table(ws, current_row, title, row_label_header,
                       row_labels, col_labels, col_keys, value_matrix,
                       row_label_map=None, section_title=None):
    """
    Write one pivot block into ws starting at current_row.
    value_matrix : dict { row_label -> { col_key -> int } }
                   values are DISTINCT counts of the Number column.
    Returns: (header_row, total_row, next_free_row)
    """
    n_cols = len(col_keys)
    total_col = n_cols + 2

    # avg across all body cells for traffic-light
    all_vals = [value_matrix.get(rl, {}).get(ck, 0)
                for rl in row_labels for ck in col_keys]
    avg_val = (sum(all_vals) / len(all_vals)) if all_vals else 0

    # Optional section heading
    if section_title:
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=total_col)
        c = ws.cell(row=current_row, column=1, value=section_title)
        c.fill = HEADER_FILL
        c.font = Font(bold=True, color=WHITE, name="Arial", size=12)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[current_row].height = 24
        current_row += 1

    # Header row
    hcell(ws, current_row, 1, row_label_header)
    for ci, cl in enumerate(col_labels, 2):
        hcell(ws, current_row, ci, cl)
    hcell(ws, current_row, total_col, "Total")
    header_row = current_row
    current_row += 1

    # Data rows
    data_start = current_row
    for pi, rl in enumerate(row_labels):
        fill_alt = ALT_PAT if pi % 2 == 0 else None
        label = row_label_map.get(rl, rl) if row_label_map else rl
        dcell(ws, current_row, 1, label, fill=fill_alt, align="left")
        row_sum = 0
        for ci, ck in enumerate(col_keys, 2):
            val = value_matrix.get(rl, {}).get(ck, 0)
            row_sum += val
            _tl_dcell(ws, current_row, ci, val, avg_val)
        fl = get_column_letter(2) + str(current_row)
        ll = get_column_letter(n_cols + 1) + str(current_row)
        tc = ws.cell(row=current_row, column=total_col, value=f"=SUM({fl}:{ll})")
        tc.font = Font(name="Arial", size=10, bold=True, color="000000")
        tc.alignment = Alignment(horizontal="center", vertical="center")
        tc.border = _border()
        tc.fill = _traffic_fill_avg(row_sum, avg_val)
        current_row += 1

    # Total row
    hcell(ws, current_row, 1, "Total", fill=TOTAL_FILL)
    for wi in range(2, n_cols + 2):
        cl = get_column_letter(wi)
        hcell(ws, current_row, wi,
              f"=SUM({cl}{data_start}:{cl}{current_row-1})", fill=TOTAL_FILL)
    tl = get_column_letter(total_col)
    hcell(ws, current_row, total_col,
          f"=SUM({tl}{data_start}:{tl}{current_row-1})", fill=TOTAL_FILL)
    total_row = current_row
    current_row += 1

    return header_row, total_row, current_row


# ── Closure data tab writer ───────────────────────────────────────────────────

def _write_closure_data(ws, df_group, start_col_name, task_col_name, log):
    if df_group.empty:
        log(" ⚠️ No rows for this group after filtering.")
        return None

    df_g = (
        df_group
        .sort_values(by=start_col_name, ascending=False, na_position="last")
        .drop_duplicates(subset=[task_col_name], keep="first")
        .copy()
    )
    before = len(df_g)
    assigned_col = find_col(df_g, ["assigned to", "assigned_to", "assignedto"])
    if assigned_col:
        df_g = df_g[df_g[assigned_col].notna() &
                    (df_g[assigned_col].astype(str).str.strip() != "")]
    removed_blank = before - len(df_g)
    log(f" {len(df_g):,} rows "
        f"(kept latest per task; dropped {removed_blank} blank Assigned To)")

    headers = list(df_g.columns)
    for ci, h in enumerate(headers, 1):
        hcell(ws, 1, ci, h)

    for ri, row in enumerate(df_g.itertuples(index=False), 2):
        fill = ALT_PAT if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(vertical="center")
            c.border = _border()
            if fill: c.fill = fill
            if headers[ci - 1] == start_col_name and val:
                c.number_format = "DD/MM/YYYY"

    autofit(ws); ws.freeze_panes = "A2"
    return df_g


# ── Closure summary table writer ──────────────────────────────────────────────

def _write_closure_summary(ws, df_clean, group_keyword,
                            task_col_name, start_col_name,
                            selected_months, log):
    if df_clean is None or df_clean.empty:
        log(" ⚠️ No data for summary table.")
        return

    assigned_col = find_col(df_clean, ["assigned to", "assigned_to", "assignedto"])
    if not assigned_col:
        log(" ⚠️ 'Assigned To' column not found — skipping summary.")
        return

    df_clean = df_clean.copy()
    df_clean["_month"] = df_clean[start_col_name].apply(
        lambda d: d.date().replace(day=1) if pd.notna(d) else None
    )
    df_sm = df_clean[df_clean["_month"].isin(selected_months)].copy()

    pivot = (
        pd.pivot_table(df_sm, index=assigned_col, columns="_month",
                       values=task_col_name, aggfunc="count", fill_value=0)
        if not df_sm.empty else pd.DataFrame()
    )

    sel_months_sorted = sorted(selected_months)
    month_labels = [fmt_month(m) for m in sel_months_sorted]
    n_months = len(sel_months_sorted)
    total_col = n_months + 2
    people = sorted(df_sm[assigned_col].dropna().unique()) if not df_sm.empty else []

    all_vals = []
    person_vals = {}
    for person in people:
        row_vals = {}
        for mo in sel_months_sorted:
            v = (int(pivot.loc[person, mo])
                 if not pivot.empty and person in pivot.index
                 and mo in pivot.columns else 0)
            row_vals[mo] = v
            all_vals.append(v)
        row_vals["_total"] = sum(row_vals[mo] for mo in sel_months_sorted)
        person_vals[person] = row_vals

    avg_val = (sum(all_vals) / len(all_vals)) if all_vals else 0

    ws.merge_cells(f"A1:{get_column_letter(total_col)}1")
    c = ws.cell(row=1, column=1,
                value=f"Closure Summary — {group_keyword} — Assigned To by Month")
    c.fill = HEADER_FILL
    c.font = Font(bold=True, color=WHITE, name="Arial", size=12)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    hcell(ws, 2, 1, "Assigned To")
    for mi, ml in enumerate(month_labels, 2):
        hcell(ws, 2, mi, ml)
    hcell(ws, 2, total_col, "Total")

    DR0 = 3
    for pi, person in enumerate(people):
        r = DR0 + pi
        row_vals = person_vals[person]
        total = row_vals["_total"]
        dcell(ws, r, 1, person, fill=None, align="left")
        for mi, mo in enumerate(sel_months_sorted, 2):
            _tl_dcell(ws, r, mi, row_vals[mo], avg_val)
        fl = get_column_letter(2) + str(r)
        ll = get_column_letter(n_months + 1) + str(r)
        tc = ws.cell(row=r, column=total_col, value=f"=SUM({fl}:{ll})")
        tc.font = Font(name="Arial", size=10, bold=True, color="000000")
        tc.alignment = Alignment(horizontal="center", vertical="center")
        tc.border = _border()
        tc.fill = _traffic_fill_avg(total, avg_val)

    TR = DR0 + len(people)
    hcell(ws, TR, 1, "Total", fill=TOTAL_FILL)
    for mi in range(2, n_months + 2):
        cl = get_column_letter(mi)
        hcell(ws, TR, mi, f"=SUM({cl}{DR0}:{cl}{TR-1})", fill=TOTAL_FILL)
    tl = get_column_letter(total_col)
    hcell(ws, TR, total_col, f"=SUM({tl}{DR0}:{tl}{TR-1})", fill=TOTAL_FILL)

    autofit(ws); ws.column_dimensions["A"].width = 32; ws.freeze_panes = "B3"

    leg_row = TR + 2
    _cell(ws, leg_row, 1, "Colour scale (avg-based, ±20% tolerance):",
          bold=True, align="left", sz=9)
    for col_i, (label, lfill) in enumerate([
            ("Below avg → Red", RED_FILL),
            ("~Average → Amber", AMBER_FILL),
            ("Above avg → Green", GREEN_FILL),
    ], 2):
        c = ws.cell(row=leg_row, column=col_i, value=label)
        c.fill = lfill; c.font = Font(name="Arial", size=9, bold=True, color="000000")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()

    log(f" ✅ {len(people)} people × {n_months} months (avg={avg_val:.1f})")


# ── Core workbook builder ─────────────────────────────────────────────────────

def build_workbook(
    raw_data_bytes: bytes,
    master_bytes: bytes,
    master_filename: str,
    selected_weeks: list,
    selected_months: list,
    log=print
) -> bytes:

    log("📂 Reading raw data...")
    df_raw = pd.read_excel(io.BytesIO(raw_data_bytes), sheet_name=0)
    log(f" {len(df_raw):,} rows | {len(df_raw.columns)} columns")

    task_col = find_col(df_raw, ["task number", "task_number", "task"])
    group_col = find_col(df_raw, ["assignment group", "assignment_group", "group"])
    start_col = find_col(df_raw, ["start", "start date", "start_date", "created"])
    priority_col = find_col(df_raw, ["priority"])
    number_col = find_col(df_raw, ["number", "no", "num"])

    if number_col:
        log(f" ✅ Number column found: '{number_col}' — all pivots will use DISTINCT COUNT of this column")
    else:
        raise ValueError(
            "Number column not found (tried: 'number', 'no', 'num').\n"
            f"Available columns: {list(df_raw.columns)}\n"
            "Please ensure your raw data has a column named 'Number' (or 'No' / 'Num')."
        )

    missing = [n for n, c in [("Task Number", task_col),
                               ("Assignment Group", group_col),
                               ("Start", start_col)] if c is None]
    if missing:
        raise ValueError(f"Required columns not found: {missing}\n"
                         f"Available: {list(df_raw.columns)}")

    df_raw = df_raw.copy()
    df_raw[start_col] = pd.to_datetime(df_raw[start_col], errors="coerce")

    # Add week / month helper columns directly on raw data
    df_raw["_Start_Week"] = df_raw[start_col].apply(
        lambda d: (d - pd.Timedelta(days=d.weekday())).date() if pd.notna(d) else None
    )
    df_raw["_Start_Month"] = df_raw[start_col].apply(
        lambda d: d.date().replace(day=1) if pd.notna(d) else None
    )

    if priority_col:
        df_raw[priority_col] = df_raw[priority_col].astype(str).str.strip()

    # ── Load master & clear managed tabs ─────────────────────────────────────
    log("📂 Loading master workbook...")
    wb = load_workbook(io.BytesIO(master_bytes))

    log("🧹 Clearing managed tabs...")
    for tab in MANAGED_TABS:
        if tab not in wb.sheetnames:
            wb.create_sheet(tab)
            log(f" Created missing sheet: '{tab}'")
        _clear_sheet(wb[tab])
        log(f" Cleared: '{tab}'")

    # ── Step 1: Original with Dups ────────────────────────────────────────────
    log("📝 Writing 'Original with Dups'...")
    ws_orig = wb["Original with Dups"]
    # Write without the helper columns (they're internal)
    df_display = df_raw.drop(columns=["_Start_Week", "_Start_Month"])
    headers = list(df_display.columns)
    for ci, h in enumerate(headers, 1):
        hcell(ws_orig, 1, ci, h)
    for ri, row in enumerate(df_display.itertuples(index=False), 2):
        fill = ALT_PAT if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            c = ws_orig.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(vertical="center"); c.border = _border()
            if fill: c.fill = fill
            if headers[ci - 1] == start_col and val:
                c.number_format = "DD/MM/YYYY"
    autofit(ws_orig); ws_orig.freeze_panes = "A2"
    log(f" ✅ {len(df_raw):,} rows written")

    # ── Filter raw data by selected weeks ─────────────────────────────────────
    df_f = df_raw[df_raw["_Start_Week"].isin(selected_weeks)].copy()
    week_labels = [fmt_week(w) for w in selected_weeks]
    n_weeks = len(selected_weeks)
    total_col_i = n_weeks + 2

    groups = sorted(df_f[group_col].dropna().unique())
    priorities = [p for p in PRIORITY_ORDER
                  if priority_col and p in df_f[priority_col].values]
    if not priorities and priority_col:
        priorities = sorted(df_f[priority_col].dropna().unique())
    log(f" {len(groups)} groups | {n_weeks} weeks selected "
        f"| {len(df_f):,} rows in week range")

    # ── Charts sheet ──────────────────────────────────────────────────────────
    ws_charts = wb["Charts"]
    _cell(ws_charts, 1, 1, "All Charts — Distinct Count of Number", bold=True, sz=14)
    ws_charts.column_dimensions["A"].width = 2
    chart_row = [3]

    def _place(chart, row_span=32):
        ws_charts.add_chart(chart, f"B{chart_row[0]}")
        chart_row[0] += row_span + 1

    # ── Step 2: WTBP (Priority × Week, distinct Number count per Group) ──────
    log("📋 Building 'WTBP' (distinct Number count)...")
    ws_wtbp = wb["WTBP"]
    current_row = 1

    for grp_idx, grp in enumerate(groups):
        df_grp = df_f[df_f[group_col] == grp]

        # Distinct count pivot: rows=priority, cols=week, values=nunique(number)
        pivot_dc = _distinct_count_pivot(df_grp, priority_col, "_Start_Week", number_col) \
                   if priority_col and not df_grp.empty else pd.DataFrame()

        vm = {}
        for prio in priorities:
            vm[prio] = {}
            for wk in selected_weeks:
                vm[prio][wk] = (int(pivot_dc.loc[prio, wk])
                                if not pivot_dc.empty and prio in pivot_dc.index
                                and wk in pivot_dc.columns else 0)

        hr, tr, current_row = _write_pivot_table(
            ws_wtbp, current_row, grp, "Priority",
            priorities, week_labels, selected_weeks, vm,
            row_label_map=PRIORITY_LABELS,
            section_title=grp
        )
        _place(_combo(ws_wtbp, hr, tr, 1, 2, n_weeks + 1, total_col_i,
                      f"{grp} – Priority by Week (Distinct #)"))
        current_row += 1
        if grp_idx < len(groups) - 1:
            current_row += 3

    autofit(ws_wtbp); ws_wtbp.column_dimensions["A"].width = 20
    log(f" ✅ {len(groups)} groups written to 'WTBP'")

    # ── Step 3: VolumeAnalysis (Group × Week, distinct Number count) ─────────
    log("📊 Building 'VolumeAnalysis' (distinct Number count)...")
    ws_va = wb["VolumeAnalysis"]

    vol_dc = _distinct_count_pivot(df_f, group_col, "_Start_Week", number_col) \
             if not df_f.empty else pd.DataFrame()

    vm_va = {}
    for grp in groups:
        vm_va[grp] = {}
        for wk in selected_weeks:
            vm_va[grp][wk] = (int(vol_dc.loc[grp, wk])
                              if not vol_dc.empty and grp in vol_dc.index
                              and wk in vol_dc.columns else 0)

    hr_va, tr_va, _ = _write_pivot_table(
        ws_va, 1, "Weekly Volume", "Assignment Group",
        groups, week_labels, selected_weeks, vm_va,
        section_title="Weekly Volume by Assignment Group (Distinct Number Count)"
    )
    autofit(ws_va); ws_va.column_dimensions["A"].width = 32; ws_va.freeze_panes = "B3"
    _place(_combo(ws_va, hr_va, tr_va, 1, 2, n_weeks + 1, n_weeks + 2,
                  "Weekly Volume – Distinct Number Count", w=28, h=16), row_span=38)
    log(" ✅ 'VolumeAnalysis' written")

    # ── Step 4: MonthlyVolume (Group × Month, distinct Number count) ─────────
    log("📅 Building 'MonthlyVolume' (distinct Number count)...")
    ws_mv = wb["MonthlyVolume"]

    df_monthly = df_raw[df_raw["_Start_Month"].isin(selected_months)].copy()
    sel_months_sorted = sorted(selected_months)
    month_labels = [fmt_month(m) for m in sel_months_sorted]
    n_months = len(sel_months_sorted)
    all_groups_mv = sorted(df_monthly[group_col].dropna().unique())

    mv_dc = _distinct_count_pivot(df_monthly, group_col, "_Start_Month", number_col) \
            if not df_monthly.empty else pd.DataFrame()

    vm_mv = {}
    for grp in all_groups_mv:
        vm_mv[grp] = {}
        for mo in sel_months_sorted:
            vm_mv[grp][mo] = (int(mv_dc.loc[grp, mo])
                              if not mv_dc.empty and grp in mv_dc.index
                              and mo in mv_dc.columns else 0)

    hr_mv, tr_mv, _ = _write_pivot_table(
        ws_mv, 1, "Monthly Volume", "Assignment Group",
        all_groups_mv, month_labels, sel_months_sorted, vm_mv,
        section_title="Monthly Volume by Assignment Group (Distinct Number Count)"
    )
    autofit(ws_mv); ws_mv.column_dimensions["A"].width = 32; ws_mv.freeze_panes = "B3"
    _place(_combo(ws_mv, hr_mv, tr_mv, 1, 2, n_months + 1, n_months + 2,
                  "Monthly Volume – Distinct Number Count", w=28, h=16), row_span=38)
    log(f" ✅ 'MonthlyVolume' written ({len(all_groups_mv)} groups × {n_months} months)")

    # ── Step 5: Closure tabs ──────────────────────────────────────────────────
    log("🔒 Building Closure tabs...")
    for keyword, data_tab, summary_tab in CLOSURE_GROUPS:
        log(f" ── {keyword} ──")
        mask = (df_raw[group_col].astype(str)
                      .str.strip().str.lower() == keyword.strip().lower())
        df_grp_raw = df_raw[mask].drop(columns=["_Start_Week", "_Start_Month"]).copy()
        log(f" {len(df_grp_raw):,} rows matched '{keyword}'")
        ws_data = wb[data_tab]
        df_clean = _write_closure_data(ws_data, df_grp_raw, start_col, task_col, log)
        ws_sum = wb[summary_tab]
        _write_closure_summary(ws_sum, df_clean, keyword, task_col, start_col,
                               selected_months, log)
    log(" ✅ All Closure tabs written")

    # ── Reorder sheets ────────────────────────────────────────────────────────
    for idx, name in enumerate(MANAGED_TABS):
        if name in wb.sheetnames:
            cur_idx = wb.sheetnames.index(name)
            wb.move_sheet(name, offset=idx - cur_idx)

    log("💾 Saving workbook...")
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    log("✅ ALL DONE!")
    return buf.read()


# ── Download helper ───────────────────────────────────────────────────────────

def _show_download(result_bytes: bytes, out_filename: str):
    from IPython.display import display, FileLink
    candidates = [os.path.expanduser("~"), tempfile.gettempdir(), os.getcwd()]
    save_dir = None
    for d in candidates:
        try:
            test = os.path.join(d, ".write_test")
            with open(test, "w") as f: f.write("x")
            os.remove(test); save_dir = d; break
        except OSError:
            continue
    if save_dir is None:
        print("❌ No writable directory found."); return
    save_path = os.path.join(save_dir, out_filename)
    with open(save_path, "wb") as f: f.write(result_bytes)
    print(f"\n✅ File saved → {save_path}\nClick the link below to download:\n")
    display(FileLink(save_path, result_html_prefix="⬇️ "))


# ── Jupyter UI ────────────────────────────────────────────────────────────────

def run_jupyter():
    import ipywidgets as widgets
    from IPython.display import display, clear_output

    display(widgets.HTML("""
        <div style="background:#1E3A5F;padding:14px 20px;border-radius:8px;
                    margin-bottom:10px">
            <h2 style="color:white;margin:0;font-family:Arial;font-size:18px">
                📊 Excel Automation Tool v12
            </h2>
            <p style="color:#aad4f5;margin:6px 0 0;font-family:Arial;font-size:12px">
                1 Upload master &nbsp;→&nbsp; 2 Upload raw data &nbsp;→&nbsp;
                3 Pick weeks &nbsp;→&nbsp; 4 Pick months &nbsp;→&nbsp;
                5 Run &nbsp;→&nbsp; 6 Download
            </p>
            <p style="color:#aad4f5;margin:4px 0 0;font-family:Arial;font-size:11px">
                Pivot values = DISTINCT COUNT of Number column · Source = Original with Dups
            </p>
        </div>
    """))

    master_uploader = widgets.FileUpload(
        accept=".xlsx", multiple=False, description="📁 Master File",
        button_style="primary", layout=widgets.Layout(width="230px"))
    master_status = widgets.HTML(
        "<span style='color:grey;font-size:12px;font-family:Arial'>"
        "The master workbook</span>")

    raw_uploader = widgets.FileUpload(
        accept=".xlsx", multiple=False, description="📁 Raw Data File",
        button_style="warning", layout=widgets.Layout(width="230px"))
    raw_status = widgets.HTML(
        "<span style='color:grey;font-size:12px;font-family:Arial'>"
        "New raw data to process</span>")

    week_header = widgets.HTML("", layout=widgets.Layout(display="none"))
    week_grid = widgets.GridBox([], layout=widgets.Layout(display="none"))
    w_sel_all_btn = widgets.Button(description="☑ Select All", button_style="info",
        layout=widgets.Layout(width="130px", display="none", margin="0 8px 8px 0"))
    w_clr_all_btn = widgets.Button(description="☐ Clear All", button_style="warning",
        layout=widgets.Layout(width="130px", display="none", margin="0 0 8px 0"))

    month_header = widgets.HTML("", layout=widgets.Layout(display="none"))
    month_grid = widgets.GridBox([], layout=widgets.Layout(display="none"))
    m_sel_all_btn = widgets.Button(description="☑ Select All", button_style="info",
        layout=widgets.Layout(width="130px", display="none", margin="0 8px 8px 0"))
    m_clr_all_btn = widgets.Button(description="☐ Clear All", button_style="warning",
        layout=widgets.Layout(width="130px", display="none", margin="0 0 8px 0"))

    run_btn = widgets.Button(description="▶ Run & Update Master",
        button_style="success",
        layout=widgets.Layout(width="230px", margin="12px 0 0 0"), disabled=True)
    out = widgets.Output(layout=widgets.Layout(
        border="1px solid #ddd", padding="12px", border_radius="6px",
        min_height="60px", margin_top="10px"))

    display(
        widgets.HTML("<b style='font-family:Arial;font-size:13px'>1 Master Workbook</b>"),
        master_uploader, master_status,
        widgets.HTML("<b style='font-family:Arial;font-size:13px'>2 Raw Data File</b>"),
        raw_uploader, raw_status,
        week_header, widgets.HBox([w_sel_all_btn, w_clr_all_btn]), week_grid,
        month_header, widgets.HBox([m_sel_all_btn, m_clr_all_btn]), month_grid,
        run_btn, out
    )

    _ctx = {"master_bytes": None, "master_filename": None, "raw_bytes": None}
    _week_cbs = {}
    _month_cbs = {}

    def _try_enable():
        run_btn.disabled = not (_ctx["master_bytes"] and _ctx["raw_bytes"]
                                and _week_cbs and _month_cbs)

    def _build_week_ui(weeks):
        _week_cbs.clear()
        boxes = []
        for w in weeks:
            cb = widgets.Checkbox(value=True, description=fmt_week(w),
                layout=widgets.Layout(width="115px"),
                style={"description_width": "0px"})
            _week_cbs[w] = cb; boxes.append(cb)
        week_header.value = (f"<b style='font-family:Arial;font-size:13px'>"
                             f"3 Select weeks ({len(weeks)} found):</b>")
        week_header.layout.display = ""
        week_grid.children = boxes
        week_grid.layout = widgets.Layout(
            grid_template_columns="repeat(6, 120px)", display="", margin="4px 0 0 0")
        w_sel_all_btn.layout.display = ""; w_clr_all_btn.layout.display = ""

    def _build_month_ui(months):
        _month_cbs.clear()
        boxes = []
        for m in months:
            cb = widgets.Checkbox(value=True, description=fmt_month(m),
                layout=widgets.Layout(width="120px"),
                style={"description_width": "0px"})
            _month_cbs[m] = cb; boxes.append(cb)
        month_header.value = (f"<b style='font-family:Arial;font-size:13px'>"
                              f"4 Select months ({len(months)} found):</b>")
        month_header.layout.display = ""
        month_grid.children = boxes
        month_grid.layout = widgets.Layout(
            grid_template_columns="repeat(4, 130px)", display="", margin="4px 0 0 0")
        m_sel_all_btn.layout.display = ""; m_clr_all_btn.layout.display = ""

    def on_master_upload(change):
        fname, content = _read_upload(master_uploader)
        if not fname: return
        _ctx["master_bytes"] = content; _ctx["master_filename"] = fname
        master_status.value = (
            f"<span style='color:green;font-size:12px;font-family:Arial'>"
            f"✅ <b>{fname}</b> loaded — {len(content)/1024:.1f} KB</span>")
        _try_enable()

    def on_raw_upload(change):
        fname, content = _read_upload(raw_uploader)
        if not fname: return
        _ctx["raw_bytes"] = content
        raw_status.value = (
            f"<span style='color:grey;font-size:12px;font-family:Arial'>"
            f"⏳ Scanning <b>{fname}</b>...</span>")
        try:
            weeks, months, _ = get_weeks_and_months_from_bytes(content)
        except Exception as e:
            raw_status.value = f"<span style='color:red;font-size:12px'>❌ Error: {e}</span>"
            return
        if not weeks:
            raw_status.value = ("<span style='color:red;font-size:12px'>"
                                "❌ No Start column found.</span>"); return
        raw_status.value = (
            f"<span style='color:green;font-size:12px;font-family:Arial'>"
            f"✅ <b>{fname}</b> — {len(content)/1024:.1f} KB — "
            f"{len(weeks)} week(s), {len(months)} month(s) detected</span>")
        _build_week_ui(weeks); _build_month_ui(months); _try_enable()

    def on_w_sel_all(_):
        for cb in _week_cbs.values(): cb.value = True
    def on_w_clr_all(_):
        for cb in _week_cbs.values(): cb.value = False
    def on_m_sel_all(_):
        for cb in _month_cbs.values(): cb.value = True
    def on_m_clr_all(_):
        for cb in _month_cbs.values(): cb.value = False

    def on_run(_):
        with out:
            clear_output()
            if not _ctx["master_bytes"]: print("❌ Upload master."); return
            if not _ctx["raw_bytes"]: print("❌ Upload raw data."); return
            sel_weeks = sorted(w for w, cb in _week_cbs.items() if cb.value)
            sel_months = sorted(m for m, cb in _month_cbs.items() if cb.value)
            if not sel_weeks: print("❌ No weeks selected."); return
            if not sel_months: print("❌ No months selected."); return
            print(f"⚙️ Weeks : {fmt_week(sel_weeks[0])} → {fmt_week(sel_weeks[-1])} "
                  f"({len(sel_weeks)} selected)\n"
                  f"⚙️ Months: {fmt_month(sel_months[0])} → {fmt_month(sel_months[-1])} "
                  f"({len(sel_months)} selected)\n")
            try:
                result_bytes = build_workbook(
                    raw_data_bytes = _ctx["raw_bytes"],
                    master_bytes = _ctx["master_bytes"],
                    master_filename = _ctx["master_filename"],
                    selected_weeks = sel_weeks,
                    selected_months = sel_months,
                    log = lambda m: print(m)
                )
                _show_download(result_bytes, _ctx["master_filename"])
            except Exception:
                print("❌ Error:"); traceback.print_exc()

    master_uploader.observe(on_master_upload, names="value")
    raw_uploader.observe(on_raw_upload, names="value")
    w_sel_all_btn.on_click(on_w_sel_all); w_clr_all_btn.on_click(on_w_clr_all)
    m_sel_all_btn.on_click(on_m_sel_all); m_clr_all_btn.on_click(on_m_clr_all)
    run_btn.on_click(on_run)


# ── CLI fallback ──────────────────────────────────────────────────────────────

def run_cli(master_path: str, raw_path: str):
    if not os.path.isfile(master_path): print(f"❌ {master_path}"); sys.exit(1)
    if not os.path.isfile(raw_path): print(f"❌ {raw_path}"); sys.exit(1)

    with open(raw_path, "rb") as f: raw_bytes = f.read()
    with open(master_path, "rb") as f: master_bytes = f.read()

    weeks, months, _ = get_weeks_and_months_from_bytes(raw_bytes)
    if not weeks: print("❌ No Start values found."); sys.exit(1)

    print(f"\n📅 {len(weeks)} weeks:")
    for i, w in enumerate(weeks): print(f" [{i:2d}] {fmt_week(w)}")
    raw = input("\nWeeks start/end index (Enter = all): ").strip()
    try:
        p = raw.split(); sel_weeks = weeks[int(p[0]):int(p[1]) + 1]
    except Exception:
        sel_weeks = weeks

    print(f"\n📆 {len(months)} months:")
    for i, m in enumerate(months): print(f" [{i:2d}] {fmt_month(m)}")
    raw = input("\nMonths start/end index (Enter = all): ").strip()
    try:
        p = raw.split(); sel_months = months[int(p[0]):int(p[1]) + 1]
    except Exception:
        sel_months = months

    result_bytes = build_workbook(
        raw_data_bytes=raw_bytes, master_bytes=master_bytes,
        master_filename=os.path.basename(master_path),
        selected_weeks=sel_weeks, selected_months=sel_months)
    with open(master_path, "wb") as f: f.write(result_bytes)
    print(f"\n💾 Updated: {master_path}")


# ── Entry ─────────────────────────────────────────────────────────────────────

def main():
    try:
        from IPython import get_ipython
        in_jupyter = get_ipython() is not None
    except ImportError:
        in_jupyter = False
    if in_jupyter:
        run_jupyter()
    else:
        if len(sys.argv) < 3:
            print("Usage: python automate_excel_v12.py <master.xlsx> <raw.xlsx>")
            sys.exit(1)
        run_cli(sys.argv[1], sys.argv[2])

main()